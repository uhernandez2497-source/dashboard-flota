#!/usr/bin/env python3
"""
actualizar_datos.py - Agente de actualizacion de datos del Dashboard
====================================================================
Descarga el Excel desde OneDrive, lo convierte a data.json y lo sube a GitHub Pages.

Uso:
  python actualizar_datos.py          # Descarga, convierte y sube a GitHub
  python actualizar_datos.py --local  # Solo descarga y convierte (sin push)

Requisitos:
  - Python 3.x
  - openpyxl (pip install openpyxl)
  - git configurado con acceso al repositorio
"""

import json
import os
import subprocess
import sys
from datetime import datetime

# ── CONFIGURACION ──
ONEDRIVE_URL = 'https://proxylogis-my.sharepoint.com/personal/carlosu_hernandez_mecanicatek_com/_layouts/15/download.aspx?share=IQBkHquuUpUmR4eUs1jTUCezAb4_AMRJLf9ZR2KC27WL7GU'
SHEET_NAME = 'Sheet1'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_JSON = os.path.join(SCRIPT_DIR, 'data.json')
EXCEL_TEMP = os.path.join(SCRIPT_DIR, '_temp_dashboard.xlsx')

# Mapeo de nombre de encabezado -> clave interna
# El script busca estas columnas por nombre, no por posicion fija
HEADER_MAP = {
    'Equipo': 'equipo',
    'Tipo Equipo': 'tipo_equipo',
    'Marca': 'marca',
    'Modelo': 'modelo',
    'Razon Reparacion': 'razon',
    'Fecha Terminacion': 'fecha_terminacion',
    'Fecha Ingreso': 'fecha_ingreso',
    'Precio Refaccion': 'refaccion',
    'Precio Mano Obra': 'mano_obra',
    'Precio Otros Talleres': 'otros',
    'Total': 'total',
    'Dias Real': 'dias_real',
    'Dias Atraso': 'dias_atraso',
    'Dias Promedio': 'dias_promedio',
    'Nodo': 'nodo',
    'Taller Orden': 'taller_orden',
    'Region': 'region',
    'Zona Cliente': 'zona_cliente',
    'Clasificacion De La Orden': 'clasificacion',
    'Tipo de Servicio': 'tipo_servicio',
    'Tiempo estandar': 'tiempo',
    'Familia': 'familia',
    'Grupo Mantenimiento': 'grupo',
    'Orden': 'orden',
    'Cliente': 'cliente',
    'Operador': 'operador',
    'Solicitante': 'solicitante',
    'Calificacion Atendido A Tiempo': 'atendido_a_tiempo',
    'Kms Acumulados': 'kms_acumulados',
    'Serie': 'serie',
}

# Columnas obligatorias (el script falla si no las encuentra)
REQUIRED_COLS = ['equipo', 'fecha_terminacion', 'total']

# Mapeo de numero de mes a nombre en español (el dashboard espera español minuscula)
MES_ESPANOL = {
    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
    5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
    9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
}

MESES_VALIDOS = set(MES_ESPANOL.values())


def log(msg):
    print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}')


def download_excel():
    """Descarga el Excel desde OneDrive usando urllib (sin dependencias externas)."""
    import urllib.request
    log('Descargando Excel desde OneDrive...')
    req = urllib.request.Request(ONEDRIVE_URL, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    })
    try:
        response = urllib.request.urlopen(req, timeout=30)
        data = response.read()
        if len(data) < 1000:
            raise Exception(f'Archivo muy pequeno ({len(data)} bytes), posible error')
        with open(EXCEL_TEMP, 'wb') as f:
            f.write(data)
        log(f'Descargado: {len(data):,} bytes -> {EXCEL_TEMP}')
        return True
    except Exception as e:
        log(f'ERROR descargando: {e}')
        return False


def detectar_columnas(header):
    """Detecta la posicion de cada columna buscando por nombre de encabezado.
    Esto hace el script resistente a cambios en el orden de columnas del Excel."""
    col = {}
    header_norm = {}
    for i, h in enumerate(header):
        if h is not None:
            header_norm[str(h).strip()] = i

    for header_name, key in HEADER_MAP.items():
        if header_name in header_norm:
            col[key] = header_norm[header_name]

    # Verificar columnas obligatorias
    missing = [k for k in REQUIRED_COLS if k not in col]
    if missing:
        log(f'  ADVERTENCIA: Columnas no encontradas en encabezado: {missing}')
        log(f'  Encabezados disponibles: {list(header_norm.keys())}')
        return None

    found = len(col)
    total = len(HEADER_MAP)
    log(f'  Columnas detectadas: {found}/{total}')
    if found < total:
        not_found = [h for h, k in HEADER_MAP.items() if k not in col]
        log(f'  No encontradas (opcionales): {not_found}')

    return col


def parse_excel():
    """Parsea el Excel y genera la lista de registros."""
    try:
        import openpyxl
    except ImportError:
        log('ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl')
        sys.exit(1)

    log(f'Parseando Excel (hoja: {SHEET_NAME})...')
    wb = openpyxl.load_workbook(EXCEL_TEMP, read_only=True, data_only=True)

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb[wb.sheetnames[0]]
        log(f'  Hoja "{SHEET_NAME}" no encontrada, usando "{ws.title}"')

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    log(f'  {len(rows)} filas leidas (incluye encabezado)')

    if not rows:
        log('  ERROR: Excel vacio')
        return []

    # Detectar columnas por nombre de encabezado
    col = detectar_columnas(rows[0])
    if col is None:
        log('  ERROR: No se pudieron detectar las columnas obligatorias')
        return []

    def safe_float(v):
        try:
            return float(v) if v is not None else 0
        except (ValueError, TypeError):
            return 0

    def safe_str(v):
        return str(v).strip() if v is not None else ''

    def get(row, key, default=None):
        """Obtiene valor de una columna por clave, o default si no existe."""
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    def safe_date_str(v):
        """Convierte fecha a string ISO para JSON."""
        if v is None:
            return ''
        try:
            if isinstance(v, datetime):
                return v.strftime('%d/%m/%Y %H:%M')
            s = str(v).strip()
            if s:
                return s
        except (ValueError, TypeError):
            pass
        return ''

    def extraer_mes_ano(row):
        """Extrae mes (español) y año desde Fecha Terminacion.
        Fuente primaria y confiable: la fecha real del registro."""
        fecha = get(row, 'fecha_terminacion')

        if fecha is not None:
            try:
                if isinstance(fecha, datetime):
                    return MES_ESPANOL[fecha.month], str(fecha.year)
                fecha_str = str(fecha).strip()
                if fecha_str and fecha_str[:4].isdigit():
                    dt = datetime.fromisoformat(fecha_str.replace(' ', 'T').split('.')[0])
                    return MES_ESPANOL[dt.month], str(dt.year)
            except (ValueError, KeyError):
                pass

        return '', ''

    records = []
    sin_fecha = 0
    for row in rows[1:]:  # Skip header
        equipo = get(row, 'equipo')
        if not equipo:
            continue

        mes, ano = extraer_mes_ano(row)
        if not mes or not ano:
            sin_fecha += 1

        records.append({
            'equipo': safe_str(equipo),
            'orden': safe_str(get(row, 'orden', '')),
            'tipo_equipo': safe_str(get(row, 'tipo_equipo', '')),
            'marca': safe_str(get(row, 'marca', '')),
            'modelo': safe_str(get(row, 'modelo', '')),
            'serie': safe_str(get(row, 'serie', '')),
            'razon_reparacion': safe_str(get(row, 'razon', '')),
            'fecha_ingreso': safe_date_str(get(row, 'fecha_ingreso')),
            'fecha_terminacion': safe_date_str(get(row, 'fecha_terminacion')),
            'precio_refaccion': safe_float(get(row, 'refaccion', 0)),
            'precio_mano_obra': safe_float(get(row, 'mano_obra', 0)),
            'precio_otros': safe_float(get(row, 'otros', 0)),
            'total': safe_float(get(row, 'total', 0)),
            'dias_real': safe_float(get(row, 'dias_real', 0)),
            'dias_atraso': safe_float(get(row, 'dias_atraso', 0)),
            'dias_promedio': safe_float(get(row, 'dias_promedio', 0)),
            'nodo': safe_str(get(row, 'nodo', '')),
            'taller_orden': safe_str(get(row, 'taller_orden', '')),
            'region': safe_str(get(row, 'region', '')),
            'zona_cliente': safe_str(get(row, 'zona_cliente', '')),
            'clasificacion': safe_str(get(row, 'clasificacion', '')),
            'tipo_servicio': safe_str(get(row, 'tipo_servicio', '')),
            'tiempo_estandar': safe_float(get(row, 'tiempo', 0)),
            'familia': safe_str(get(row, 'familia', '')),
            'mes': mes,
            'ano': ano,
            'grupo_manto': safe_str(get(row, 'grupo', '')),
            'cliente': safe_str(get(row, 'cliente', '')),
            'operador': safe_str(get(row, 'operador', '')),
            'solicitante': safe_str(get(row, 'solicitante', '')),
            'atendido_a_tiempo': safe_str(get(row, 'atendido_a_tiempo', '')),
            'kms_acumulados': safe_float(get(row, 'kms_acumulados', 0)),
        })

    log(f'  {len(records)} registros validos extraidos')
    if sin_fecha:
        log(f'  ADVERTENCIA: {sin_fecha} registros sin fecha de terminacion')
    return records


def validar_datos(records):
    """Valida la integridad de los datos extraidos. Retorna True si pasan, False si hay errores criticos."""
    ok = True

    # Validar que hay registros
    if len(records) == 0:
        log('VALIDACION FALLO: 0 registros extraidos')
        return False

    # Validar meses
    meses_invalidos = set()
    anos_invalidos = set()
    sin_mes = 0
    sin_ano = 0
    for r in records:
        if not r['mes']:
            sin_mes += 1
        elif r['mes'] not in MESES_VALIDOS:
            meses_invalidos.add(r['mes'])
        if not r['ano']:
            sin_ano += 1
        elif not (r['ano'].isdigit() and 2020 <= int(r['ano']) <= 2040):
            anos_invalidos.add(r['ano'])

    if meses_invalidos:
        log(f'VALIDACION FALLO: Meses invalidos encontrados: {meses_invalidos}')
        ok = False
    if anos_invalidos:
        log(f'VALIDACION FALLO: Anos invalidos encontrados: {anos_invalidos}')
        ok = False
    if sin_mes > 0:
        log(f'VALIDACION ADVERTENCIA: {sin_mes} registros sin mes')
    if sin_ano > 0:
        log(f'VALIDACION ADVERTENCIA: {sin_ano} registros sin ano')

    # Resumen de distribucion
    combos = {}
    for r in records:
        key = f'{r["mes"]} {r["ano"]}'
        combos[key] = combos.get(key, 0) + 1
    log(f'  Distribucion: {dict(sorted(combos.items()))}')

    # Comparar con data.json anterior (detectar caidas de registros)
    if os.path.exists(OUTPUT_JSON):
        try:
            with open(OUTPUT_JSON, 'r', encoding='utf-8') as f:
                prev = json.load(f)
            prev_count = prev.get('count', 0)
            if prev_count > 0 and len(records) < prev_count * 0.8:
                log(f'VALIDACION FALLO: Registros cayeron de {prev_count} a {len(records)} (>20% menos)')
                ok = False
            elif prev_count > 0 and len(records) < prev_count:
                log(f'VALIDACION ADVERTENCIA: Registros bajaron de {prev_count} a {len(records)}')
        except (json.JSONDecodeError, KeyError):
            pass

    if ok:
        log('  Validacion OK')
    return ok


def save_json(records):
    """Guarda los registros como data.json y data.js."""
    output = {
        'updated': datetime.now().isoformat(),
        'count': len(records),
        'data': records
    }
    # data.json (para GitHub Pages y dashboard original)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

    size_kb = os.path.getsize(OUTPUT_JSON) / 1024
    log(f'Guardado: {OUTPUT_JSON} ({size_kb:.0f} KB, {len(records)} registros)')

    # data.js (para carga directa en dashboard_ordenes.html via file://)
    output_js = os.path.join(SCRIPT_DIR, 'data.js')
    json_str = json.dumps(output, ensure_ascii=False, separators=(',', ':'))
    with open(output_js, 'w', encoding='utf-8') as f:
        f.write(f'window.__FLEET_DATA__={json_str};')

    size_kb_js = os.path.getsize(output_js) / 1024
    log(f'Guardado: {output_js} ({size_kb_js:.0f} KB)')


def git_push():
    """Commit y push a GitHub."""
    log('Subiendo a GitHub...')
    try:
        # Configurar git si estamos en GitHub Actions
        if os.getenv('GITHUB_ACTIONS'):
            subprocess.run(['git', 'config', 'user.name', os.getenv('GIT_AUTHOR_NAME', 'GitHub Actions')],
                         cwd=SCRIPT_DIR, check=True, capture_output=True)
            subprocess.run(['git', 'config', 'user.email', os.getenv('GIT_AUTHOR_EMAIL', 'actions@github.com')],
                         cwd=SCRIPT_DIR, check=True, capture_output=True)

        subprocess.run(['git', 'add', 'data.json', 'data.js'], cwd=SCRIPT_DIR, check=True, capture_output=True)
        ts = datetime.now().strftime('%Y-%m-%d %H:%M')
        result = subprocess.run(
            ['git', 'commit', '-m', f'Actualizar datos: {ts}'],
            cwd=SCRIPT_DIR, capture_output=True, text=True
        )
        if result.returncode != 0:
            if 'nothing to commit' in result.stdout or 'nothing to commit' in result.stderr:
                log('Sin cambios en los datos - no hay nada que subir')
                return True
            log(f'ERROR en commit: {result.stderr}')
            return False

        # Determinar la rama principal (puede ser master o main)
        branch_result = subprocess.run(
            ['git', 'rev-parse', '--abbrev-ref', 'HEAD'],
            cwd=SCRIPT_DIR, capture_output=True, text=True
        )
        branch = branch_result.stdout.strip() if branch_result.returncode == 0 else 'master'

        result = subprocess.run(
            ['git', 'push', 'origin', branch],
            cwd=SCRIPT_DIR, capture_output=True, text=True
        )
        if result.returncode != 0:
            log(f'ERROR en push: {result.stderr}')
            return False

        log(f'Push exitoso a GitHub Pages (rama: {branch})')
        return True
    except FileNotFoundError:
        log('ERROR: git no encontrado en PATH')
        return False


def cleanup():
    """Limpia archivos temporales."""
    if os.path.exists(EXCEL_TEMP):
        os.remove(EXCEL_TEMP)


def main():
    local_only = '--local' in sys.argv

    print('=' * 50)
    print('  AGENTE DE ACTUALIZACION - Dashboard Flota')
    print('=' * 50)
    print()

    # Paso 1: Descargar Excel
    if not download_excel():
        log('FALLO: No se pudo descargar el Excel')
        cleanup()
        sys.exit(1)

    # Paso 2: Parsear Excel
    records = parse_excel()
    if not records:
        log('FALLO: No se encontraron registros')
        cleanup()
        sys.exit(1)

    # Paso 3: Validar datos
    if not validar_datos(records):
        log('FALLO: Los datos no pasaron la validacion')
        cleanup()
        sys.exit(1)

    # Paso 4: Generar JSON
    save_json(records)

    # Paso 5: Subir a GitHub (si no es --local)
    if not local_only:
        git_push()
    else:
        log('Modo local: no se sube a GitHub (usa sin --local para push)')

    # Limpieza
    cleanup()

    print()
    log('COMPLETADO')
    print(f'  Registros: {len(records)}')
    print(f'  Archivo:   {OUTPUT_JSON}')
    if not local_only:
        print(f'  GitHub:    https://uhernandez2497-source.github.io/dashboard-flota/')
    print()


if __name__ == '__main__':
    main()
